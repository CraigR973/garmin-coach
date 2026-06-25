"""Spreadsheet import and baseline helpers for historical sleep backfill."""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta
from pathlib import Path
from statistics import mean, median, pstdev, quantiles
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from sqlalchemy import Select, select
from sqlalchemy.ext.asyncio import AsyncSession

from src.models.coaching import DailyMetric, MetricBaseline, Sleep
from src.models.profile import Profile

SLEEP_HISTORY_SOURCE = "sleep_history_xlsx"
SPO2_HRV_RELIABLE_FROM = date(2026, 6, 11)
EXCEL_EPOCH = datetime(1899, 12, 30)
MIDDAY = time(12, 0)


@dataclass(frozen=True)
class SleepHistoryRow:
    calendar_date: date
    score: int | None
    age_adjusted_score: int | None
    resting_heart_rate_bpm: int | None
    body_battery_charge: int | None
    average_spo2_pct: float | None
    average_respiration: float | None
    skin_temp_change_c: float | None
    hrv_7_day_avg_ms: int | None
    qualifier: str | None
    sleep_need_sec: int | None
    sleep_start_utc: datetime | None
    sleep_end_utc: datetime | None
    duration_sec: int | None
    raw_payload: dict[str, Any]

    def to_sleep_fields(self) -> dict[str, Any]:
        return {
            "calendar_date": self.calendar_date,
            "sleep_start_utc": self.sleep_start_utc,
            "sleep_end_utc": self.sleep_end_utc,
            "score": self.score,
            "age_adjusted_score": self.age_adjusted_score,
            "qualifier": self.qualifier,
            "duration_sec": self.duration_sec,
            "average_spo2_pct": self.average_spo2_pct,
            "average_respiration": self.average_respiration,
            "resting_heart_rate_bpm": self.resting_heart_rate_bpm,
            "body_battery_change": self.body_battery_charge,
            "factors_json": {
                "source": SLEEP_HISTORY_SOURCE,
                "sleepNeedSec": self.sleep_need_sec,
                "skinTempChangeC": self.skin_temp_change_c,
                "qualifier": self.qualifier,
            },
            "raw_payload": self.raw_payload,
        }

    def to_daily_metric_fields(self) -> dict[str, Any]:
        return {
            "calendar_date": self.calendar_date,
            "recorded_at_utc": self.sleep_end_utc,
            "resting_heart_rate_bpm": self.resting_heart_rate_bpm,
            "body_battery_charged": self.body_battery_charge,
            "hrv_weekly_avg_ms": self.hrv_7_day_avg_ms,
            "raw_payload": self.raw_payload,
        }


@dataclass(frozen=True)
class BaselineMetricSpec:
    metric_key: str
    metric_label: str
    reliability_start_date: date | None = None


BASELINE_SPECS: tuple[BaselineMetricSpec, ...] = (
    BaselineMetricSpec("sleep_score", "Sleep score"),
    BaselineMetricSpec("age_adjusted_sleep_score", "Age-adjusted sleep score"),
    BaselineMetricSpec("resting_heart_rate_bpm", "Resting heart rate"),
    BaselineMetricSpec("body_battery_charge", "Body Battery charge"),
    BaselineMetricSpec(
        "average_spo2_pct",
        "Average Pulse Ox",
        reliability_start_date=SPO2_HRV_RELIABLE_FROM,
    ),
    BaselineMetricSpec("average_respiration", "Average respiration"),
    BaselineMetricSpec(
        "hrv_7_day_avg_ms",
        "7-day average HRV",
        reliability_start_date=SPO2_HRV_RELIABLE_FROM,
    ),
)


@dataclass(frozen=True)
class BaselineSample:
    """One calendar day's metric values keyed by baseline ``metric_key``.

    The shared currency between the xlsx importer and the DB-history backfill
    (``services/metric_baselines.py``) so both feed the *same*
    :func:`compute_metric_baselines` core.
    """

    calendar_date: date
    values: Mapping[str, float | int | None]


@dataclass(frozen=True)
class SleepHistoryImportResult:
    rows_parsed: int
    rows_skipped: int
    dry_run: bool
    sleep_created: int = 0
    sleep_updated: int = 0
    daily_metrics_created: int = 0
    daily_metrics_updated: int = 0
    baselines_created: int = 0
    baselines_updated: int = 0


def parse_sleep_history_workbook(
    path: str | Path, timezone_name: str
) -> tuple[list[SleepHistoryRow], int]:
    workbook = load_workbook(filename=Path(path), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[SleepHistoryRow] = []
    skipped = 0

    for values in sheet.iter_rows(min_row=2, values_only=True):
        row_map = {
            str(headers[index]): value for index, value in enumerate(values) if index < len(headers)
        }
        parsed = _parse_sleep_history_row(row_map, timezone_name)
        if parsed is None:
            skipped += 1
            continue
        rows.append(parsed)

    workbook.close()
    return rows, skipped


def compute_metric_baselines(
    samples: Sequence[BaselineSample],
    *,
    source: str,
    specs: Sequence[BaselineMetricSpec] = BASELINE_SPECS,
) -> list[dict[str, Any]]:
    """Summarise per-day samples into persisted ``metric_baselines`` field dicts.

    Pure + source-agnostic: both the 84-night xlsx importer and the DB-history
    backfill feed :class:`BaselineSample`s through this one core so the stats —
    and the #45 SpO2/HRV reliability cutoff (rows before
    ``spec.reliability_start_date`` are dropped from that metric and surfaced as
    ``excluded_sample_count``) — are computed identically. ``source`` is stamped
    onto each row so provenance stays honest and the
    ``(user_id, metric_key, source)`` unique key never collides across origins.
    """
    if not samples:
        return []

    window_start = min(sample.calendar_date for sample in samples)
    window_end = max(sample.calendar_date for sample in samples)
    baselines: list[dict[str, Any]] = []

    for spec in specs:
        values = [
            float(value)
            for sample in samples
            if (value := sample.values.get(spec.metric_key)) is not None
            and (
                spec.reliability_start_date is None
                or sample.calendar_date >= spec.reliability_start_date
            )
        ]
        excluded_count = sum(
            1
            for sample in samples
            if sample.values.get(spec.metric_key) is not None
            and spec.reliability_start_date is not None
            and sample.calendar_date < spec.reliability_start_date
        )
        if not values:
            continue

        ordered = sorted(values)
        lower_quartile, upper_quartile = _quartiles(ordered)
        baselines.append(
            {
                "metric_key": spec.metric_key,
                "metric_label": spec.metric_label,
                "source": source,
                "window_start_date": window_start,
                "window_end_date": window_end,
                "reliability_start_date": spec.reliability_start_date,
                "sample_count": len(ordered),
                "excluded_sample_count": excluded_count,
                "mean_value": mean(ordered),
                "median_value": median(ordered),
                "min_value": ordered[0],
                "max_value": ordered[-1],
                "lower_quartile_value": lower_quartile,
                "upper_quartile_value": upper_quartile,
                "stddev_value": pstdev(ordered) if len(ordered) > 1 else 0.0,
                "raw_payload": {
                    "metricKey": spec.metric_key,
                    "metricLabel": spec.metric_label,
                    "source": source,
                    "windowStartDate": window_start.isoformat(),
                    "windowEndDate": window_end.isoformat(),
                    "reliabilityStartDate": (
                        spec.reliability_start_date.isoformat()
                        if spec.reliability_start_date is not None
                        else None
                    ),
                    "sampleCount": len(ordered),
                    "excludedSampleCount": excluded_count,
                    "values": ordered,
                },
            }
        )

    return baselines


def _sleep_history_sample(row: SleepHistoryRow) -> BaselineSample:
    return BaselineSample(
        calendar_date=row.calendar_date,
        values={
            "sleep_score": row.score,
            "age_adjusted_sleep_score": row.age_adjusted_score,
            "resting_heart_rate_bpm": row.resting_heart_rate_bpm,
            "body_battery_charge": row.body_battery_charge,
            "average_spo2_pct": row.average_spo2_pct,
            "average_respiration": row.average_respiration,
            "hrv_7_day_avg_ms": row.hrv_7_day_avg_ms,
        },
    )


def build_metric_baselines(rows: list[SleepHistoryRow]) -> list[dict[str, Any]]:
    return compute_metric_baselines(
        [_sleep_history_sample(row) for row in rows],
        source=SLEEP_HISTORY_SOURCE,
    )


class SleepHistoryImportService:
    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def import_workbook(
        self,
        profile: Profile,
        workbook_path: str | Path,
        *,
        dry_run: bool = False,
    ) -> SleepHistoryImportResult:
        rows, skipped = parse_sleep_history_workbook(workbook_path, profile.timezone)
        if not rows:
            return SleepHistoryImportResult(rows_parsed=0, rows_skipped=skipped, dry_run=dry_run)

        existing_sleep = await self._load_existing_sleep(profile.id, rows)
        existing_metrics = await self._load_existing_metrics(profile.id, rows)
        existing_baselines = await self._load_existing_baselines(profile.id)

        sleep_created = 0
        sleep_updated = 0
        daily_created = 0
        daily_updated = 0
        baseline_created = 0
        baseline_updated = 0

        for row in rows:
            sleep = existing_sleep.get(row.calendar_date)
            if sleep is None:
                sleep_created += 1
                if not dry_run:
                    sleep = Sleep(user_id=profile.id, calendar_date=row.calendar_date)
                    self.session.add(sleep)
                    existing_sleep[row.calendar_date] = sleep
                    _apply_fields(sleep, row.to_sleep_fields())
            else:
                changed = _apply_fields(sleep, row.to_sleep_fields(), apply=not dry_run)
                if changed:
                    sleep_updated += 1

            metric = existing_metrics.get(row.calendar_date)
            if metric is None:
                daily_created += 1
                if not dry_run:
                    metric = DailyMetric(user_id=profile.id, calendar_date=row.calendar_date)
                    self.session.add(metric)
                    existing_metrics[row.calendar_date] = metric
                    _apply_fields(metric, row.to_daily_metric_fields())
            else:
                changed = _apply_fields(metric, row.to_daily_metric_fields(), apply=not dry_run)
                if changed:
                    daily_updated += 1

        for baseline_fields in build_metric_baselines(rows):
            key = str(baseline_fields["metric_key"])
            baseline = existing_baselines.get(key)
            if baseline is None:
                baseline_created += 1
                if not dry_run:
                    baseline = MetricBaseline(user_id=profile.id, metric_key=key)
                    self.session.add(baseline)
                    existing_baselines[key] = baseline
                    _apply_fields(baseline, baseline_fields)
            else:
                changed = _apply_fields(baseline, baseline_fields, apply=not dry_run)
                if changed:
                    baseline_updated += 1

        if not dry_run:
            await self.session.commit()

        return SleepHistoryImportResult(
            rows_parsed=len(rows),
            rows_skipped=skipped,
            dry_run=dry_run,
            sleep_created=sleep_created,
            sleep_updated=sleep_updated,
            daily_metrics_created=daily_created,
            daily_metrics_updated=daily_updated,
            baselines_created=baseline_created,
            baselines_updated=baseline_updated,
        )

    async def _load_existing_sleep(
        self, user_id: Any, rows: list[SleepHistoryRow]
    ) -> dict[date, Sleep]:
        statement = _date_window_select(
            select(Sleep), Sleep.user_id == user_id, Sleep.calendar_date, rows
        )
        result = await self.session.execute(statement)
        return {row.calendar_date: row for row in result.scalars().all()}

    async def _load_existing_metrics(
        self, user_id: Any, rows: list[SleepHistoryRow]
    ) -> dict[date, DailyMetric]:
        statement = _date_window_select(
            select(DailyMetric), DailyMetric.user_id == user_id, DailyMetric.calendar_date, rows
        )
        result = await self.session.execute(statement)
        return {row.calendar_date: row for row in result.scalars().all()}

    async def _load_existing_baselines(self, user_id: Any) -> dict[str, MetricBaseline]:
        result = await self.session.execute(
            select(MetricBaseline).where(
                MetricBaseline.user_id == user_id,
                MetricBaseline.source == SLEEP_HISTORY_SOURCE,
            )
        )
        return {row.metric_key: row for row in result.scalars().all()}


def _date_window_select(
    statement: Select[Any], user_filter: Any, date_column: Any, rows: list[SleepHistoryRow]
) -> Select[Any]:
    return statement.where(
        user_filter,
        date_column >= min(row.calendar_date for row in rows),
        date_column <= max(row.calendar_date for row in rows),
    )


def _apply_fields(instance: Any, fields: dict[str, Any], *, apply: bool = True) -> bool:
    changed = False
    for key, value in fields.items():
        if getattr(instance, key) != value:
            changed = True
            if apply:
                setattr(instance, key, value)
    return changed


def _parse_sleep_history_row(row_map: dict[str, Any], timezone_name: str) -> SleepHistoryRow | None:
    calendar_date = _parse_excel_date(row_map.get("Sleep Score 4 Weeks"))
    if calendar_date is None:
        return None

    sleep_end_utc = _local_time_to_utc(
        calendar_date,
        _parse_excel_time(row_map.get("Wake Time")),
        timezone_name,
    )
    sleep_start_utc = _local_time_to_utc(
        calendar_date,
        _parse_excel_time(row_map.get("Bedtime")),
        timezone_name,
        bedtime=True,
    )

    duration_sec: int | None = None
    if (
        sleep_start_utc is not None
        and sleep_end_utc is not None
        and sleep_end_utc > sleep_start_utc
    ):
        duration_sec = int((sleep_end_utc - sleep_start_utc).total_seconds())

    raw_payload = {
        "source": SLEEP_HISTORY_SOURCE,
        "date": calendar_date.isoformat(),
        "score": _json_value(row_map.get("Score")),
        "restingHeartRate": _json_value(row_map.get("Resting Heart Rate")),
        "bodyBatteryCharge": _json_value(row_map.get("Body Battery Charge")),
        "pulseOx": _json_value(row_map.get("Pulse Ox")),
        "respiration": _json_value(row_map.get("Respiration")),
        "skinTempChange": _json_value(row_map.get("Skin Temp Change")),
        "sevenDayAverageHrv": _json_value(row_map.get("7 Day Average HRV ")),
        "quality": _json_value(row_map.get("Quality")),
        "duration": _json_value(row_map.get("Duration")),
        "sleepNeed": _json_value(row_map.get("Sleep Need")),
        "bedtime": _json_value(row_map.get("Bedtime")),
        "wakeTime": _json_value(row_map.get("Wake Time")),
    }

    score = _to_int(row_map.get("Score"))
    return SleepHistoryRow(
        calendar_date=calendar_date,
        score=score,
        age_adjusted_score=min(score + 4, 100) if score is not None else None,
        resting_heart_rate_bpm=_to_int(row_map.get("Resting Heart Rate")),
        body_battery_charge=_to_int(row_map.get("Body Battery Charge")),
        average_spo2_pct=_to_float(row_map.get("Pulse Ox")),
        average_respiration=_to_float(row_map.get("Respiration")),
        skin_temp_change_c=_parse_signed_float(row_map.get("Skin Temp Change")),
        hrv_7_day_avg_ms=_to_int(row_map.get("7 Day Average HRV ")),
        qualifier=_to_upper_str(row_map.get("Quality")),
        sleep_need_sec=_parse_duration_text(row_map.get("Sleep Need")),
        sleep_start_utc=sleep_start_utc,
        sleep_end_utc=sleep_end_utc,
        duration_sec=duration_sec,
        raw_payload=raw_payload,
    )


def _parse_excel_date(value: Any) -> date | None:
    if value in (None, "", "--"):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    serial = _to_float(value)
    if serial is None:
        return None
    return (EXCEL_EPOCH + timedelta(days=serial)).date()


def _parse_excel_time(value: Any) -> time | None:
    if value in (None, "", "--"):
        return None
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    serial = _to_float(value)
    if serial is None:
        return None
    seconds = int(round((serial % 1) * 24 * 60 * 60))
    seconds %= 24 * 60 * 60
    return (datetime.min + timedelta(seconds=seconds)).time()


def _local_time_to_utc(
    calendar_date: date,
    local_time: time | None,
    timezone_name: str,
    *,
    bedtime: bool = False,
) -> datetime | None:
    if local_time is None:
        return None
    local_date = calendar_date
    if bedtime and local_time >= MIDDAY:
        local_date = calendar_date - timedelta(days=1)
    zone = ZoneInfo(timezone_name)
    local_dt = datetime.combine(local_date, local_time, tzinfo=zone)
    return local_dt.astimezone(UTC).replace(tzinfo=None)


def _parse_duration_text(value: Any) -> int | None:
    if value in (None, "", "--"):
        return None
    if isinstance(value, (int, float)):
        return int(float(value) * 24 * 60 * 60)
    text = str(value).strip().lower()
    if "h" not in text and "min" not in text:
        return None
    hours = 0
    minutes = 0
    if "h" in text:
        hours_text, _, remainder = text.partition("h")
        hours = int(hours_text.strip() or "0")
        text = remainder
    if "min" in text:
        minutes_text, _, _ = text.partition("min")
        minutes = int(minutes_text.strip() or "0")
    return (hours * 60 + minutes) * 60


def _quartiles(values: list[float]) -> tuple[float, float]:
    if len(values) == 1:
        return values[0], values[0]
    cuts = quantiles(values, n=4, method="inclusive")
    return cuts[0], cuts[2]


def _parse_signed_float(value: Any) -> float | None:
    if value in (None, "", "--"):
        return None
    return _to_float(str(value).replace("°", ""))


def _to_int(value: Any) -> int | None:
    if value in (None, "", "--"):
        return None
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return None


def _to_float(value: Any) -> float | None:
    if value in (None, "", "--"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_upper_str(value: Any) -> str | None:
    if value in (None, "", "--"):
        return None
    text = str(value).strip()
    return text.upper() if text else None


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return value
